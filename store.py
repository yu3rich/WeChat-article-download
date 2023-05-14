import os

from openpyxl import Workbook  # 表格处理
from openpyxl import load_workbook


class Dataxls:
    # 设置文件位置
    def __init__(self, xlsname, header):
        self.xlsname = xlsname
        self.header = header
        openxlsx = os.path.exists(self.xlsname)  # 检测表是否存在
        if openxlsx:
            self.score_book = load_workbook(self.xlsname)
            self.score_sheet = self.score_book.active
            print('表格存在，即将续写')
        else:
            self.score_book = Workbook()  # 创建文件设置目录
            self.score_sheet = self.score_book.create_sheet('Sheet1', 0)  # 设置工作表名
            # 设置表头
            self.score_sheet.append(self.header)
            #score_sheet.save(self.xlsname)

    # 第一种模式 逐行写入
    def xls_1(self, list):
        self.score_sheet.append(list)
        self.score_book.save(self.xlsname)   # 报讯表格
